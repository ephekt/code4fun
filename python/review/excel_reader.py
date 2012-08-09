# -*- coding: utf-8 -*-
from openpyxl.reader.excel import load_workbook
import sys
import glob
import MySQLdb
from multiprocessing import Pool

def remove_non_ascii(s): return "".join(i for i in s if ord(i)<128)

def extract_row(row):
	s = []
	for cell in row:
		v = cell.internal_value
		if type(v) in (int, float):
			v = str(v)
		elif v is None:
			v = "NULL"
		else:
			v = "%s" % dbconn.escape(str(remove_non_ascii(v)))
		s.append(v)
	return s

def run ( workbook_name ):
	print("Working on workbook: %s" % workbook_name)
	
	#dbconn   = MySQLdb.connect("remotedb.connection.com","username","password123",'Data')
	dbconn   = MySQLdb.connect("localhost","root",'','Data')
	dbcursor = dbconn.cursor(MySQLdb.cursors.Cursor)
	wb       = load_workbook(filename = workbook_name, use_iterators = True)
	
	for worksheet_name in wb.get_sheet_names():
		print("Processing %s from %s" % (worksheet_name,workbook_name))
		
		ws = wb.get_sheet_by_name(name = worksheet_name) # ws is now an IterableWorksheet
		# total_rows = ws.get_highest_row()
		
		headers = None
		for row in ws.iter_rows(): # it brings a new method: iter_rows()
			if not headers:
				headers = extra_row(row)
			else:
				try:
					qstr = "insert into data (%s) values (%s)" % (','.join(headers), ','.join(extract_row(row)))
					dbcursor.execute(qstr)
					dbconn.commit()
				except Exception, e:
					print v
					print query_string
					print e

if __name__ == '__main__':
	workbooks = glob.glob('/Users/miker/optisol/src/ruby/adhoc/*.xlsx')
	p = Pool(8)
	p.map(run,workbooks)

